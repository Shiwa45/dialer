"""
Advanced Analytics Reports - Phase 4.3

This module provides comprehensive analytics and reporting:
1. Historical trend analysis
2. Agent performance reports
3. Campaign ROI tracking
4. Export functionality (Excel/PDF)
"""

import logging
from datetime import datetime, timedelta
from typing import Dict, List, Optional
from io import BytesIO

from django.utils import timezone
from django.db.models import Avg, Count, Sum, Q, F, FloatField
from django.db.models.functions import TruncDate, TruncHour, TruncWeek, Cast
from django.http import HttpResponse

logger = logging.getLogger(__name__)


class AnalyticsEngine:
    """
    Advanced Analytics Engine
    
    Phase 4.3: Comprehensive reporting and analytics
    
    Usage:
        engine = AnalyticsEngine()
        trends = engine.get_call_trends(campaign_id, days=30)
        comparison = engine.compare_periods(campaign_id, period1, period2)
    """
    
    def get_call_trends(
        self, 
        campaign_id: int = None, 
        days: int = 30,
        granularity: str = 'day'
    ) -> Dict:
        """
        Get call volume trends over time
        
        Args:
            campaign_id: Optional campaign filter
            days: Number of days to analyze
            granularity: 'hour', 'day', or 'week'
        
        Returns:
            dict: Trend data with labels and values
        """
        from calls.models import CallLog
        
        end_date = timezone.now()
        start_date = end_date - timedelta(days=days)
        
        calls = CallLog.objects.filter(start_time__gte=start_date)
        
        if campaign_id:
            calls = calls.filter(campaign_id=campaign_id)
        
        # Choose truncation function
        if granularity == 'hour':
            trunc_func = TruncHour('start_time')
        elif granularity == 'week':
            trunc_func = TruncWeek('start_time')
        else:
            trunc_func = TruncDate('start_time')
        
        # Aggregate by period
        trend_data = calls.annotate(
            period=trunc_func
        ).values('period').annotate(
            total_calls=Count('id'),
            answered_calls=Count('id', filter=Q(answer_time__isnull=False)),
            sales=Count('id', filter=Q(disposition__is_sale=True)),
            avg_duration=Avg('talk_duration'),
            total_duration=Sum('talk_duration')
        ).order_by('period')
        
        # Format response
        labels = []
        totals = []
        answered = []
        sales = []
        
        for item in trend_data:
            if item['period']:
                if granularity == 'hour':
                    labels.append(item['period'].strftime('%m/%d %H:00'))
                elif granularity == 'week':
                    labels.append(f"Week {item['period'].strftime('%W')}")
                else:
                    labels.append(item['period'].strftime('%m/%d'))
                
                totals.append(item['total_calls'])
                answered.append(item['answered_calls'])
                sales.append(item['sales'])
        
        return {
            'labels': labels,
            'datasets': {
                'total_calls': totals,
                'answered_calls': answered,
                'sales': sales
            },
            'summary': {
                'total_calls': sum(totals),
                'answered_calls': sum(answered),
                'sales': sum(sales),
                'avg_daily_calls': sum(totals) / max(len(totals), 1)
            }
        }
    
    def get_hourly_performance(self, campaign_id: int = None, days: int = 7) -> Dict:
        """
        Get average performance by hour of day
        
        Helps identify best times to call.
        """
        from calls.models import CallLog
        from django.db.models.functions import ExtractHour
        
        end_date = timezone.now()
        start_date = end_date - timedelta(days=days)
        
        calls = CallLog.objects.filter(start_time__gte=start_date)
        
        if campaign_id:
            calls = calls.filter(campaign_id=campaign_id)
        
        hourly = calls.annotate(
            hour=ExtractHour('start_time')
        ).values('hour').annotate(
            total_calls=Count('id'),
            answered_calls=Count('id', filter=Q(answer_time__isnull=False)),
            sales=Count('id', filter=Q(disposition__is_sale=True)),
        ).order_by('hour')
        
        # Calculate rates
        result = []
        for item in hourly:
            total = item['total_calls']
            answered = item['answered_calls']
            sales_count = item['sales']
            
            contact_rate = (answered / total * 100) if total > 0 else 0
            conversion_rate = (sales_count / answered * 100) if answered > 0 else 0
            
            result.append({
                'hour': item['hour'],
                'hour_formatted': f"{item['hour']:02d}:00",
                'total_calls': total,
                'answered_calls': answered,
                'sales': sales_count,
                'contact_rate': round(contact_rate, 1),
                'conversion_rate': round(conversion_rate, 1)
            })
        
        # Find best hours
        best_contact = max(result, key=lambda x: x['contact_rate']) if result else None
        best_conversion = max(result, key=lambda x: x['conversion_rate']) if result else None
        
        return {
            'hourly_data': result,
            'best_contact_hour': best_contact['hour'] if best_contact else None,
            'best_conversion_hour': best_conversion['hour'] if best_conversion else None
        }
    
    def compare_periods(
        self,
        campaign_id: int,
        period1_start: datetime,
        period1_end: datetime,
        period2_start: datetime,
        period2_end: datetime
    ) -> Dict:
        """
        Compare metrics between two time periods
        
        Args:
            campaign_id: Campaign to analyze
            period1_*: First period dates
            period2_*: Second period dates
        
        Returns:
            dict: Comparison metrics with changes
        """
        from calls.models import CallLog
        
        def get_period_metrics(start, end):
            calls = CallLog.objects.filter(
                campaign_id=campaign_id,
                start_time__gte=start,
                start_time__lte=end
            )
            
            total = calls.count()
            answered = calls.filter(answer_time__isnull=False).count()
            sales = calls.filter(disposition__is_sale=True).count()
            
            agg = calls.aggregate(
                total_duration=Sum('talk_duration'),
                avg_duration=Avg('talk_duration')
            )
            
            contact_rate = (answered / total * 100) if total > 0 else 0
            conversion_rate = (sales / answered * 100) if answered > 0 else 0
            
            return {
                'total_calls': total,
                'answered_calls': answered,
                'sales': sales,
                'contact_rate': round(contact_rate, 1),
                'conversion_rate': round(conversion_rate, 1),
                'total_duration': agg['total_duration'] or 0,
                'avg_duration': round(agg['avg_duration'] or 0, 1)
            }
        
        period1_metrics = get_period_metrics(period1_start, period1_end)
        period2_metrics = get_period_metrics(period2_start, period2_end)
        
        # Calculate changes
        changes = {}
        for key in period1_metrics:
            p1 = period1_metrics[key]
            p2 = period2_metrics[key]
            
            if p1 and p1 != 0:
                change_pct = ((p2 - p1) / p1) * 100
            else:
                change_pct = 0 if p2 == 0 else 100
            
            changes[key] = {
                'period1': p1,
                'period2': p2,
                'change': p2 - p1,
                'change_percent': round(change_pct, 1)
            }
        
        return {
            'period1': {
                'start': period1_start.isoformat(),
                'end': period1_end.isoformat(),
                'metrics': period1_metrics
            },
            'period2': {
                'start': period2_start.isoformat(),
                'end': period2_end.isoformat(),
                'metrics': period2_metrics
            },
            'changes': changes
        }
    
    def get_agent_leaderboard(
        self,
        campaign_id: int = None,
        days: int = 7,
        metric: str = 'sales'
    ) -> List[Dict]:
        """
        Get agent leaderboard ranked by specified metric
        
        Args:
            campaign_id: Optional campaign filter
            days: Time period
            metric: Ranking metric (sales, calls, conversion, quality)
        
        Returns:
            list: Ranked agent list
        """
        from calls.models import CallLog
        from django.contrib.auth.models import User
        
        end_date = timezone.now()
        start_date = end_date - timedelta(days=days)
        
        calls = CallLog.objects.filter(
            start_time__gte=start_date,
            agent__isnull=False
        )
        
        if campaign_id:
            calls = calls.filter(campaign_id=campaign_id)
        
        # Aggregate by agent
        agent_stats = calls.values(
            'agent_id',
            'agent__username',
            'agent__first_name',
            'agent__last_name'
        ).annotate(
            total_calls=Count('id'),
            answered_calls=Count('id', filter=Q(answer_time__isnull=False)),
            sales=Count('id', filter=Q(disposition__is_sale=True)),
            total_duration=Sum('talk_duration'),
            avg_quality=Avg('quality_score')
        )
        
        # Calculate rates and format
        leaderboard = []
        for agent in agent_stats:
            total = agent['total_calls']
            answered = agent['answered_calls']
            sales_count = agent['sales']
            
            contact_rate = (answered / total * 100) if total > 0 else 0
            conversion_rate = (sales_count / answered * 100) if answered > 0 else 0
            
            name = f"{agent['agent__first_name'] or ''} {agent['agent__last_name'] or ''}".strip()
            if not name:
                name = agent['agent__username']
            
            leaderboard.append({
                'agent_id': agent['agent_id'],
                'name': name,
                'total_calls': total,
                'answered_calls': answered,
                'sales': sales_count,
                'contact_rate': round(contact_rate, 1),
                'conversion_rate': round(conversion_rate, 1),
                'total_duration': agent['total_duration'] or 0,
                'avg_quality': round(agent['avg_quality'] or 0, 1)
            })
        
        # Sort by specified metric
        sort_key = {
            'sales': lambda x: x['sales'],
            'calls': lambda x: x['total_calls'],
            'conversion': lambda x: x['conversion_rate'],
            'quality': lambda x: x['avg_quality'],
            'contact': lambda x: x['contact_rate']
        }.get(metric, lambda x: x['sales'])
        
        leaderboard.sort(key=sort_key, reverse=True)
        
        # Add ranks
        for idx, agent in enumerate(leaderboard):
            agent['rank'] = idx + 1
        
        return leaderboard
    
    def get_campaign_roi(self, campaign_id: int, days: int = 30) -> Dict:
        """
        Calculate campaign ROI metrics
        
        Args:
            campaign_id: Campaign to analyze
            days: Time period
        
        Returns:
            dict: ROI metrics
        """
        from calls.models import CallLog
        from campaigns.models import Campaign
        from leads.models import Lead
        
        campaign = Campaign.objects.filter(id=campaign_id).first()
        if not campaign:
            return {'error': 'Campaign not found'}
        
        end_date = timezone.now()
        start_date = end_date - timedelta(days=days)
        
        calls = CallLog.objects.filter(
            campaign_id=campaign_id,
            start_time__gte=start_date
        )
        
        # Get call metrics
        total_calls = calls.count()
        answered_calls = calls.filter(answer_time__isnull=False).count()
        sales = calls.filter(disposition__is_sale=True).count()
        
        # Get talk time
        total_duration = calls.aggregate(total=Sum('talk_duration'))['total'] or 0
        
        # Estimate costs (these would come from campaign settings)
        cost_per_minute = getattr(campaign, 'cost_per_minute', 0.03)
        revenue_per_sale = getattr(campaign, 'revenue_per_sale', 100)
        agent_cost_per_hour = getattr(campaign, 'agent_hourly_cost', 15)
        
        # Calculate costs
        telecom_cost = (total_duration / 60) * cost_per_minute
        agent_hours = total_duration / 3600
        agent_cost = agent_hours * agent_cost_per_hour
        total_cost = telecom_cost + agent_cost
        
        # Calculate revenue
        total_revenue = sales * revenue_per_sale
        
        # Calculate ROI
        profit = total_revenue - total_cost
        roi_percent = (profit / total_cost * 100) if total_cost > 0 else 0
        
        # Cost per metrics
        cost_per_call = total_cost / total_calls if total_calls > 0 else 0
        cost_per_contact = total_cost / answered_calls if answered_calls > 0 else 0
        cost_per_sale = total_cost / sales if sales > 0 else 0
        
        return {
            'campaign_id': campaign_id,
            'campaign_name': campaign.name,
            'period_days': days,
            'metrics': {
                'total_calls': total_calls,
                'answered_calls': answered_calls,
                'sales': sales,
                'talk_hours': round(agent_hours, 1)
            },
            'costs': {
                'telecom_cost': round(telecom_cost, 2),
                'agent_cost': round(agent_cost, 2),
                'total_cost': round(total_cost, 2),
                'cost_per_call': round(cost_per_call, 2),
                'cost_per_contact': round(cost_per_contact, 2),
                'cost_per_sale': round(cost_per_sale, 2)
            },
            'revenue': {
                'total_revenue': round(total_revenue, 2),
                'revenue_per_sale': revenue_per_sale
            },
            'roi': {
                'profit': round(profit, 2),
                'roi_percent': round(roi_percent, 1)
            }
        }


class ReportExporter:
    """
    Export reports to various formats
    
    Supports: Excel (.xlsx), CSV, PDF
    """
    
    def export_to_excel(self, data: Dict, report_name: str) -> BytesIO:
        """
        Export data to Excel file
        
        Args:
            data: Report data
            report_name: Name for the report
        
        Returns:
            BytesIO: Excel file buffer
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils.dataframe import dataframe_to_rows
        except ImportError:
            raise ImportError("openpyxl is required for Excel export")
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = report_name[:31]  # Excel limit
        
        # Style
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        
        # Write data based on structure
        row = 1
        
        if 'summary' in data:
            ws.cell(row=row, column=1, value='Summary').font = Font(bold=True, size=14)
            row += 1
            
            for key, value in data['summary'].items():
                ws.cell(row=row, column=1, value=key.replace('_', ' ').title())
                ws.cell(row=row, column=2, value=value)
                row += 1
            
            row += 1
        
        if 'labels' in data and 'datasets' in data:
            # Time series data
            headers = ['Date'] + list(data['datasets'].keys())
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header.replace('_', ' ').title())
                cell.font = header_font
                cell.fill = header_fill
            
            row += 1
            
            for idx, label in enumerate(data['labels']):
                ws.cell(row=row, column=1, value=label)
                for col, (key, values) in enumerate(data['datasets'].items(), 2):
                    ws.cell(row=row, column=col, value=values[idx] if idx < len(values) else '')
                row += 1
        
        if isinstance(data, list):
            # List data (like leaderboard)
            if data:
                headers = list(data[0].keys())
                
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=col, value=header.replace('_', ' ').title())
                    cell.font = header_font
                    cell.fill = header_fill
                
                row += 1
                
                for item in data:
                    for col, key in enumerate(headers, 1):
                        ws.cell(row=row, column=col, value=item.get(key, ''))
                    row += 1
        
        # Auto-adjust column widths
        for column_cells in ws.columns:
            length = max(len(str(cell.value or '')) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 50)
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer
    
    def export_to_csv(self, data: List[Dict]) -> BytesIO:
        """
        Export data to CSV
        
        Args:
            data: List of dictionaries
        
        Returns:
            BytesIO: CSV file buffer
        """
        import csv
        
        buffer = BytesIO()
        
        if not data:
            return buffer
        
        # Get headers from first row
        headers = list(data[0].keys())
        
        # Write CSV
        writer = csv.DictWriter(
            buffer, 
            fieldnames=headers,
            extrasaction='ignore'
        )
        
        # Write as text, then encode
        import io
        text_buffer = io.StringIO()
        writer = csv.DictWriter(text_buffer, fieldnames=headers)
        writer.writeheader()
        writer.writerows(data)
        
        buffer.write(text_buffer.getvalue().encode('utf-8'))
        buffer.seek(0)
        
        return buffer
    
    def get_http_response(
        self, 
        buffer: BytesIO, 
        filename: str, 
        content_type: str
    ) -> HttpResponse:
        """
        Create HTTP response for file download
        
        Args:
            buffer: File buffer
            filename: Download filename
            content_type: MIME type
        
        Returns:
            HttpResponse: File download response
        """
        response = HttpResponse(
            buffer.getvalue(),
            content_type=content_type
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


# ============================================================================
# Report Views
# ============================================================================

"""
Add to reports/views.py:

from reports.analytics import AnalyticsEngine, ReportExporter

@login_required
@supervisor_required
def analytics_dashboard(request):
    '''Advanced analytics dashboard'''
    engine = AnalyticsEngine()
    
    campaign_id = request.GET.get('campaign')
    days = int(request.GET.get('days', 30))
    
    context = {
        'trends': engine.get_call_trends(campaign_id, days),
        'hourly': engine.get_hourly_performance(campaign_id, days),
        'leaderboard': engine.get_agent_leaderboard(campaign_id, days)[:10],
    }
    
    if campaign_id:
        context['roi'] = engine.get_campaign_roi(int(campaign_id), days)
    
    return render(request, 'reports/analytics_dashboard.html', context)


@login_required
@supervisor_required
def export_report(request):
    '''Export report to file'''
    engine = AnalyticsEngine()
    exporter = ReportExporter()
    
    report_type = request.GET.get('type', 'trends')
    format = request.GET.get('format', 'xlsx')
    campaign_id = request.GET.get('campaign')
    days = int(request.GET.get('days', 30))
    
    # Get data based on report type
    if report_type == 'trends':
        data = engine.get_call_trends(campaign_id, days)
    elif report_type == 'leaderboard':
        data = engine.get_agent_leaderboard(campaign_id, days)
    elif report_type == 'hourly':
        data = engine.get_hourly_performance(campaign_id, days)
    elif report_type == 'roi' and campaign_id:
        data = engine.get_campaign_roi(int(campaign_id), days)
    else:
        return HttpResponse('Invalid report type', status=400)
    
    # Export
    filename = f"{report_type}_report_{timezone.now().strftime('%Y%m%d')}"
    
    if format == 'xlsx':
        buffer = exporter.export_to_excel(data, report_type)
        return exporter.get_http_response(
            buffer,
            f"{filename}.xlsx",
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    elif format == 'csv':
        if isinstance(data, list):
            buffer = exporter.export_to_csv(data)
        else:
            return HttpResponse('CSV export not supported for this report', status=400)
        return exporter.get_http_response(buffer, f"{filename}.csv", 'text/csv')
    
    return HttpResponse('Invalid format', status=400)
"""
